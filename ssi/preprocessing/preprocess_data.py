from typing import List, Tuple, Dict, Optional
from openpyxl import load_workbook
from ..data_logging import DataLogger
from .files import get_revenue_files_in_folder
from ..constants import Constants
import pandas as pd
import os
import tqdm


def convert_ah_receipts(input_file, coicop_sheet_prefix: str = "coi") -> pd.DataFrame:
    ah_workbook = load_workbook(filename=input_file)
    ah_sheet_names = ah_workbook.sheetnames
    coicop_sheets = [sheet for sheet in ah_sheet_names if sheet.lower(
    ).startswith(coicop_sheet_prefix)]

    all_receipts_df = pd.DataFrame()
    for sheet_name in coicop_sheets:
        ah_receipts_df = pd.read_excel(input_file, sheet_name=sheet_name)
        ah_receipts_df = ah_receipts_df.rename(columns={'Kassabonomschrijving': Constants.RECEIPT_TEXT_COLUMN,
                                                        'ArtikelEAN': Constants.PRODUCT_ID_COLUMN,
                                                        'IsbaOmschrijving': 'isba_description',
                                                        'Isba': 'isba_number',
                                                        'Esba': 'esba_number',
                                                        'BG': Constants.STORE_ID_COLUMN,
                                                        'Coicop': Constants.COICOP_LABEL_COLUMN})
        ah_receipts_df[Constants.PRODUCT_ID_COLUMN] = ah_receipts_df[Constants.PRODUCT_ID_COLUMN].astype(
            str)
        ah_receipts_df.isba_number = ah_receipts_df.isba_number.astype(str)
        ah_receipts_df.esba_number = ah_receipts_df.esba_number.astype(str)
        ah_receipts_df[Constants.STORE_ID_COLUMN] = ah_receipts_df[Constants.STORE_ID_COLUMN].astype(
            str)

        ah_receipts_df = ah_receipts_df.rename(
            columns={column_name: column_name.lower() for column_name in ah_receipts_df.columns})
        all_receipts_df = pd.concat([all_receipts_df, ah_receipts_df])

    return all_receipts_df


def year_week_to_date(year_week_column: pd.Series) -> pd.Series:
    return pd.to_datetime(year_week_column.astype(str) + '0', format="%Y%W%w")


def convert_jumbo_receipts(input_file,
                           delimiter: str = "|",
                           year_month_column: str = Constants.YEAR_MONTH_COLUMN,
                           encoding: str = "latin1"
                           ) -> pd.DataFrame:
    jumbo_receipts_df = pd.read_csv(
        input_file, sep=delimiter, encoding=encoding)

    print("separator", delimiter, "encoding", encoding)
    print(jumbo_receipts_df.head())

    jumbo_receipts_df = jumbo_receipts_df.rename(columns={'NUM_ISO_JAARWEEK': 'year_week',
                                                          'NUM_VESTIGING': 'branch_id',
                                                          'NUM_EAN': Constants.PRODUCT_ID_COLUMN,
                                                          'CBS_EAN': Constants.PRODUCT_ID_COLUMN,
                                                          'NUM_ARTIKEL': 'article_number',
                                                          'NAM_ARTIKEL': Constants.RECEIPT_TEXT_COLUMN,
                                                          })
    # Convert year-week to year-month
    jumbo_receipts_df[year_month_column] = year_week_to_date(
        jumbo_receipts_df["year_week"]).dt.strftime('%Y%m')
    jumbo_receipts_df[Constants.PRODUCT_ID_COLUMN] = jumbo_receipts_df[Constants.PRODUCT_ID_COLUMN].astype(
        str)

    return jumbo_receipts_df


def split_coicop(coicop_column: pd.Series) -> pd.DataFrame:
    return pd.DataFrame({column: coicop_column.str[:index + 2]
                        for index, column in enumerate(Constants.COICOP_LEVELS_COLUMNS)
                         })


def split_month_year(month_year_string: str) -> Tuple[str, str]:
    return month_year_string[:4], month_year_string[4:]


def split_month_year_column(dataframe: pd.DataFrame, month_year_column: str = "month") -> pd.DataFrame:
    dataframe["year"], dataframe["month"] = zip(
        *dataframe[month_year_column].apply(split_month_year))
    return dataframe


def add_leading_zero(dataframe: pd.DataFrame, coicop_column: str = Constants.COICOP_LABEL_COLUMN) -> pd.DataFrame:
    shorter_columns = dataframe[coicop_column].str.len() == 5
    dataframe.loc[shorter_columns, coicop_column] = dataframe[shorter_columns][coicop_column].apply(
        lambda s: f"0{s}")
    return dataframe


def add_coicop_levels(dataframe: pd.DataFrame, coicop_column: str = Constants.COICOP_LABEL_COLUMN) -> pd.DataFrame:
    unique_coicop = pd.Series(
        dataframe[dataframe[coicop_column].str.len() == 6][coicop_column].unique())
    split_coicop_df = split_coicop(unique_coicop)
    return dataframe.merge(split_coicop_df, on=coicop_column, suffixes=['', '_y'])


def get_category_counts(dataframe: pd.DataFrame, coicop_column: str, product_id_column: str) -> pd.DataFrame:
    return dataframe.groupby(by=[coicop_column])[product_id_column].nunique(
    ).reset_index().rename(columns={product_id_column: "count"})


def add_unique_product_id(dataframe: pd.DataFrame, column_name: str = "product_id", product_description_column: str = "ean_name") -> pd.DataFrame:
    dataframe[column_name] = dataframe[product_description_column].apply(
        hash)
    return dataframe


def filter_columns(dataframe: pd.DataFrame, columns: Optional[List[str]]) -> pd.DataFrame:
    return dataframe if not columns else dataframe[columns]


def rename_columns(dataframe: pd.DataFrame, column_mapping: dict) -> pd.DataFrame:
    return dataframe.rename(columns=column_mapping)


def add_store_names(dataframe: pd.DataFrame,
                    store_names: Dict[str, str],
                    store_id_column: str,
                    store_name_column: str) -> pd.DataFrame:
    dataframe[store_name_column] = dataframe[store_id_column].map(
        store_names)
    return dataframe


def preprocess_data(dataframe: pd.DataFrame,
                    columns: List[str],
                    coicop_column: str,
                    product_description_column: str,
                    column_mapping: Dict[str, str],
                    store_name_mapping: Dict[str, str]
                    ) -> pd.DataFrame:
    # Lidl file is losing records here
    # 1367 records have COICOP < 5 (length 1) and are filtered out
    dataframe = filter_columns(dataframe, columns)
    dataframe = rename_columns(
        dataframe, column_mapping)
    dataframe = add_leading_zero(dataframe, coicop_column=coicop_column)
    dataframe = split_month_year_column(
        dataframe, month_year_column=Constants.YEAR_MONTH_COLUMN)
    dataframe = add_coicop_levels(dataframe, coicop_column=coicop_column)
    dataframe = add_store_names(dataframe, store_name_mapping,
                                Constants.STORE_ID_COLUMN, Constants.STORE_NAME_COLUMN)

    return dataframe


def combine_revenue_files(revenue_files: List[str],
                          sort_columns: List[str],
                          sort_order: List[bool], engine: str = "pyarrow") -> pd.DataFrame:
    combined_dataframe = pd.concat([pd.read_parquet(revenue_file, engine=engine)
                                    for revenue_file in tqdm.tqdm(revenue_files)])
    combined_dataframe = combined_dataframe.sort_values(
        by=sort_columns, ascending=sort_order).reset_index(drop=True)
    return combined_dataframe


def combine_revenue_files_in_folder(data_directory: str, supermarket_name: str, sort_columns: List[str], sort_order: List[bool], filename_prefix: str = "Omzet") -> pd.DataFrame:
    revenue_files = get_revenue_files_in_folder(
        data_directory, supermarket_name, filename_prefix)
    return combine_revenue_files(revenue_files, sort_columns=sort_columns, sort_order=sort_order)


def save_combined_revenue_files(data_directory: str,
                                output_filename: str,
                                supermarket_name: str,
                                log_directory: str,
                                sort_columns: Dict[str, bool],
                                selected_columns: List[str],
                                coicop_level_columns: List[str],
                                column_mapping: Dict[str, str],
                                coicop_column: str = Constants.COICOP_LABEL_COLUMN,
                                product_description_column: str = "ean_name",
                                filename_prefix: str = "Omzet",
                                engine: str = "pyarrow"):

    supermarket_log_directory = os.path.join(log_directory, supermarket_name)
    data_logger = DataLogger(supermarket_log_directory)

    combined_df = combine_revenue_files_in_folder(
        data_directory, supermarket_name, sort_columns=list(sort_columns.keys()), sort_order=list(sort_columns.values()), filename_prefix=filename_prefix)
    data_logger.log_before_preprocessing(combined_df, coicop_column)

    combined_df = preprocess_data(
        combined_df,
        columns=selected_columns,
        coicop_column=coicop_column,
        product_description_column=product_description_column,
        column_mapping=column_mapping
    )

    data_logger.log_after_preprocessing(
        combined_df, coicop_column, coicop_level_columns, product_description_column)

    combined_df.to_parquet(os.path.join(
        data_directory, output_filename), engine=engine)
